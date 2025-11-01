import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from pathlib import Path
import json

class ExcelTools:
    """Excel operations ke liye tools"""
    
    def __init__(self, file_path=None):
        self.file_path = file_path
        self.df = None
        self.wb = None
        
    def read_excel(self, file_path, sheet_name=0):
        """Excel file read karo"""
        try:
            self.file_path = file_path
            self.df = pd.read_excel(file_path, sheet_name=sheet_name)
            return f"✅ File read successful! Shape: {self.df.shape}\n\nFirst 5 rows:\n{self.df.head()}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def get_data_info(self):
        """Data ki information"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        info = {
            "rows": len(self.df),
            "columns": len(self.df.columns),
            "column_names": list(self.df.columns),
            "data_types": self.df.dtypes.to_dict()
        }
        return json.dumps(info, indent=2, default=str)
    
    def create_excel(self, data_dict, output_path, sheet_name="Sheet1"):
        """Naya Excel file banao"""
        try:
            df = pd.DataFrame(data_dict)
            df.to_excel(output_path, sheet_name=sheet_name, index=False)
            return f"✅ File created: {output_path}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def add_data(self, data_dict, output_path):
        """Data add karo Excel me"""
        try:
            df = pd.DataFrame(data_dict)
            df.to_excel(output_path, index=False)
            return f"✅ Data added to {output_path}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def filter_data(self, column, value):
        """Data filter karo"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        try:
            filtered_df = self.df[self.df[column] == value]
            return f"Filtered Results:\n{filtered_df}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def calculate_sum(self, column):
        """Column ka sum nikalo"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        try:
            total = self.df[column].sum()
            return f"Sum of {column}: {total}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def calculate_average(self, column):
        """Column ka average nikalo"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        try:
            avg = self.df[column].mean()
            return f"Average of {column}: {avg}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def sort_data(self, column, ascending=True):
        """Data sort karo"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        try:
            self.df = self.df.sort_values(by=column, ascending=ascending)
            return f"✅ Data sorted by {column}\n{self.df.head()}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def add_column(self, column_name, values):
        """Naya column add karo"""
        if self.df is None:
            return "❌ Pehle file read karo!"
        
        try:
            self.df[column_name] = values
            return f"✅ Column '{column_name}' added!"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def save_excel(self, output_path):
        """Modified Excel save karo"""
        if self.df is None:
            return "❌ No data to save!"
        
        try:
            self.df.to_excel(output_path, index=False)
            return f"✅ File saved: {output_path}"
        except Exception as e:
            return f"❌ Error: {str(e)}"
    
    def format_excel(self, file_path, output_path):
        """Excel file ko format karo (colors, fonts)"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Header formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(output_path)
            return f"✅ File formatted and saved: {output_path}"
        except Exception as e:
            return f"❌ Error: {str(e)}"

# Test function
if __name__ == "__main__":
    print("✅ Excel Tools loaded successfully!")